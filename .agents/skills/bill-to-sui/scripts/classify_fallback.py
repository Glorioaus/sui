"""
classify_fallback.py —— LLM 兜底分类的桥梁脚本

设计原则（脚本主导，LLM 增强）：
- 脚本不直接调用 LLM API（不绑死 SDK）
- 分两阶段：extract 生成待分类清单 → LLM（skill 运行时）填充 → apply 回写 xlsx
- LLM 不可用时，extract 的结果不被填充，原始 xlsx 不受影响（确定性路径永远可用）

工作流：
    1. python classify_fallback.py extract <merged.xlsx> <tasks.json>
       → 提取落到默认兜底分类的交易，输出待分类清单
    2. （LLM 在 skill 运行时读取 tasks.json，对每条给出 category/subcategory，
       写入 results.json）
    3. python classify_fallback.py apply <merged.xlsx> <results.json>
       → 把 LLM 结果回写到 xlsx，生成 _enriched.xlsx

判定需要增强的条件（与 run_pipeline.mark_needs_enrichment 一致）：
    - 支出落到「其他杂项/其他支出」
    - 收入落到「其他收入/意外来钱」
"""
import argparse
import json
import os
import sys
from typing import List, Tuple

try:
    import openpyxl
except ImportError:
    print("错误：需要 openpyxl，请 pip install openpyxl", file=sys.stderr)
    sys.exit(2)


# 各 sheet 的列索引（1-based，与 excel_generator.py 一致）
# 支出/收入：A交易类型 B日期 C分类 D子分类 E账户 F金额 G成员 H商家 I项目 J备注
# 转账：A交易类型 B日期 C转出 D转入 E金额 F成员 G商家 H项目 I备注
SHEET_LAYOUT = {
    "支出": {"category_col": 3, "subcategory_col": 4, "account_col": 5, "amount_col": 6, "desc_col": 10},
    "收入": {"category_col": 3, "subcategory_col": 4, "account_col": 5, "amount_col": 6, "desc_col": 10},
    "转账": {"category_col": None, "subcategory_col": None, "account_col": 3, "amount_col": 5, "desc_col": 9},
}


def needs_enrichment(sheet: str, category: str, subcategory: str) -> bool:
    """判断一行是否需要 LLM 兜底分类"""
    if sheet == "支出":
        return category == "其他杂项" and subcategory == "其他支出"
    if sheet == "收入":
        return category == "其他收入" and subcategory == "意外来钱"
    return False


def extract_tasks(xlsx_path: str) -> List[dict]:
    """遍历 xlsx 三个 sheet，提取需要增强的行"""
    wb = openpyxl.load_workbook(xlsx_path)
    tasks = []
    task_id = 0

    for sheet_name in ["支出", "收入", "转账"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        layout = SHEET_LAYOUT[sheet_name]

        for row_idx in range(2, ws.max_row + 1):
            # 跳过空行
            if ws.cell(row=row_idx, column=1).value is None:
                continue

            category = ""
            subcategory = ""
            if layout["category_col"]:
                category = str(ws.cell(row=row_idx, column=layout["category_col"]).value or "")
                subcategory = str(ws.cell(row=row_idx, column=layout["subcategory_col"]).value or "")

            if not needs_enrichment(sheet_name, category, subcategory):
                continue

            amount = ws.cell(row=row_idx, column=layout["amount_col"]).value
            account = str(ws.cell(row=row_idx, column=layout["account_col"]).value or "")
            desc = str(ws.cell(row=row_idx, column=layout["desc_col"]).value or "")
            merchant_col = 8 if sheet_name in ("支出", "收入") else 7
            merchant = str(ws.cell(row=row_idx, column=merchant_col).value or "")

            tasks.append({
                "id": task_id,
                "sheet": sheet_name,
                "row": row_idx,
                "transaction_type": sheet_name.rstrip("出").replace("收", "收入") if sheet_name == "收入" else sheet_name,
                "date": str(ws.cell(row=row_idx, column=2).value or ""),
                "description": desc,
                "merchant": merchant,
                "amount": float(amount) if amount is not None else 0.0,
                "account": account,
                "current_category": category,
                "current_subcategory": subcategory,
            })
            task_id += 1

    wb.close()
    return tasks


def cmd_extract(args):
    tasks = extract_tasks(args.merged_xlsx)
    payload = {
        "source_xlsx": os.path.basename(args.merged_xlsx),
        "extracted_at": __import__("datetime").datetime.now().isoformat(timespec="seconds"),
        "instruction": (
            "对每条 task，根据 description/merchant/amount/account 判断最合适的 category 和 subcategory。"
            "参考 references/classification.md 的分类体系。"
            "若无法判断，confidence 写 'low' 并保持原 current_category/current_subcategory。"
            "输出格式见 results.json。"
        ),
        "available_categories": {
            "支出": [
                "食品酒水-早午晚餐", "食品酒水-饮料", "食品酒水-买菜", "食品酒水-水果零食",
                "衣服饰品-衣服裤子", "衣服饰品-鞋帽包包",
                "居家物业-日常用品", "居家物业-超市", "居家物业-水电煤气", "居家物业-维修保养",
                "行车交通-公共交通", "行车交通-打车租车", "行车交通-停车费", "行车交通-油费",
                "交流通讯-手机费", "交流通讯-上网费",
                "休闲娱乐-休闲玩乐", "休闲娱乐-运动健身", "休闲娱乐-会员", "休闲娱乐-旅游度假",
                "学习进修-书报杂志", "学习进修-数码装备",
                "人情往来-送礼请客", "人情往来-孝敬家长",
                "医疗保健-药品费", "医疗保健-治疗费",
                "金融保险-银行手续", "金融保险-保险",
                "其他杂项-其他支出", "其他杂项-张颖支出",
            ],
            "收入": [
                "职业收入-工资收入", "职业收入-利息收入", "职业收入-奖金收入", "职业收入-兼职收入",
                "理财-股票", "理财-基金",
                "其他收入-退款", "其他收入-报销", "其他收入-抢红包", "其他收入-礼金收入",
                "其他收入-意外来钱", "其他收入-经营所得",
            ],
        },
        "tasks": tasks,
    }
    with open(args.tasks_json, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"提取完成：{len(tasks)} 条待分类 → {args.tasks_json}")


def cmd_apply(args):
    """把 LLM 填充的 results.json 回写到 xlsx"""
    with open(args.results_json, "r", encoding="utf-8") as f:
        payload = json.load(f)
    results = payload.get("results", [])

    wb = openpyxl.load_workbook(args.merged_xlsx)
    applied = 0
    skipped = 0

    for r in results:
        sheet = r.get("sheet")
        row = r.get("row")
        category = r.get("category")
        subcategory = r.get("subcategory")
        confidence = r.get("confidence", "high")

        if not sheet or not row or not category:
            skipped += 1
            continue
        if sheet not in wb.sheetnames:
            skipped += 1
            continue
        # 低置信度不改写（保留原分类）
        if confidence == "low":
            skipped += 1
            continue

        ws = wb[sheet]
        layout = SHEET_LAYOUT.get(sheet)
        if not layout or not layout["category_col"]:
            skipped += 1
            continue

        ws.cell(row=row, column=layout["category_col"]).value = category
        ws.cell(row=row, column=layout["subcategory_col"]).value = subcategory or ""
        applied += 1

    out_path = args.output or args.merged_xlsx.replace(".xlsx", "_enriched.xlsx")
    wb.save(out_path)
    wb.close()
    print(f"应用完成：改写 {applied} 条，跳过 {skipped} 条 → {out_path}")


def main():
    parser = argparse.ArgumentParser(description="LLM 兜底分类桥梁")
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_extract = sub.add_parser("extract", help="从 merged xlsx 提取待分类清单")
    p_extract.add_argument("merged_xlsx")
    p_extract.add_argument("tasks_json")
    p_extract.set_defaults(func=cmd_extract)

    p_apply = sub.add_parser("apply", help="把 LLM 结果回写到 xlsx")
    p_apply.add_argument("merged_xlsx")
    p_apply.add_argument("results_json")
    p_apply.add_argument("--output", "-o", help="输出文件（默认 *_enriched.xlsx）")
    p_apply.set_defaults(func=cmd_apply)

    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
