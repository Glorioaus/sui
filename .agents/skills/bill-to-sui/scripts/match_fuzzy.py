"""
match_fuzzy.py —— LLM 模糊匹配的桥梁脚本

处理 merge.py 留下的两类模糊场景（详见 transfer-rules.md）：
1. 转账目标为通用"信用卡"：identify_transfers 第二轮匹配失败
2. 亲属卡未精确匹配：process_family_card 未在银行卡找到日期+金额完全匹配的交易

设计原则（与 classify_fallback.py 一致）：
- 脚本不直接调 LLM
- 两阶段：extract 生成待匹配清单 + 候选池 → LLM 判断 → apply 回写
- LLM 不可用时，原始结果不受影响

注意：match_fuzzy 主要针对"已合并但留有模糊标记"的中间结果。
最干净的工作流是：run_pipeline --emit-json → 读 transactions.json 找模糊条目 →
LLM 判断 → 直接改 transactions.json → 重新生成 xlsx。
但为了不改 src/ 的 ExcelGenerator，本脚本走"读取 xlsx → 改写 xlsx"的路径。
"""
import argparse
import json
import os
import sys
from typing import List

try:
    import openpyxl
except ImportError:
    print("错误：需要 openpyxl", file=sys.stderr)
    sys.exit(2)


def extract_unmatched_transfers(xlsx_path: str) -> List[dict]:
    """
    从转账 sheet 提取 transfer_to_account 为"信用卡"的行（merge.py 第二轮未匹配）。
    """
    wb = openpyxl.load_workbook(xlsx_path)
    tasks = []
    if "转账" not in wb.sheetnames:
        wb.close()
        return tasks

    ws = wb["转账"]
    tid = 0
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value is None:
            continue
        to_account = str(ws.cell(row=row, column=4).value or "")  # D 转入账户
        if to_account != "信用卡":
            continue

        tasks.append({
            "id": tid,
            "kind": "transfer_target",
            "sheet": "转账",
            "row": row,
            "date": str(ws.cell(row=row, column=2).value or ""),
            "from_account": str(ws.cell(row=row, column=3).value or ""),
            "to_account": to_account,
            "amount": float(ws.cell(row=row, column=5).value or 0),
            "description": str(ws.cell(row=row, column=9).value or ""),
        })
        tid += 1

    wb.close()
    return tasks


def extract_unmatched_family_cards(xlsx_path: str) -> List[dict]:
    """
    从支出 sheet 提取 subcategory 含"支出"且 category="其他杂项"的亲属消费候选。
    （process_family_card 匹配成功后会把银行卡交易重分类为"其他杂项-{使用者}支出"，
     这是已匹配的；这里提取的是潜在未匹配的亲属卡标记——但标记在 merge 阶段已被清理，
     所以实际上家属卡未匹配场景较少，留作扩展。）
    """
    # 当前 merge.py 会在最终输出前清理所有 __FAMILY_CARD__ 标记，
    # 未匹配的亲属卡会转为正式交易（其他杂项-{使用者}支出）。
    # 因此这里暂不提取，留作后续扩展点。
    return []


def cmd_extract(args):
    transfers = extract_unmatched_transfers(args.merged_xlsx)
    family = extract_unmatched_family_cards(args.merged_xlsx)

    candidate_accounts = {
        "credit_cards": ["中信信用卡", "浦发信用卡", "招商信用卡", "建行信用卡"],
        "wallets": ["微信", "支付宝", "余额宝", "花呗", "京东白条"],
    }

    payload = {
        "source_xlsx": os.path.basename(args.merged_xlsx),
        "instruction": (
            "对每条 transfer_target 任务，根据 description/from_account/amount/date 判断转入账户。"
            "优先匹配候选 credit_cards/wallets 中的具体账户名。"
            "若信息不足以判断，confidence='low'，keep_as='信用卡'（保持原值）。"
            "输出 results 数组，每条含 id/sheet/row/to_account/confidence。"
        ),
        "candidate_accounts": candidate_accounts,
        "tasks": transfers + family,
    }

    with open(args.tasks_json, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"提取完成：{len(transfers)} 条未匹配转账，{len(family)} 条亲属卡候选 → {args.tasks_json}")


def cmd_apply(args):
    with open(args.results_json, "r", encoding="utf-8") as f:
        payload = json.load(f)
    results = payload.get("results", [])

    wb = openpyxl.load_workbook(args.merged_xlsx)
    applied = 0
    skipped = 0

    for r in results:
        sheet = r.get("sheet")
        row = r.get("row")
        confidence = r.get("confidence", "high")
        to_account = r.get("to_account")
        keep_as = r.get("keep_as")

        if confidence == "low" or not to_account:
            skipped += 1
            continue
        if sheet != "转账" or sheet not in wb.sheetnames:
            skipped += 1
            continue

        ws = wb[sheet]
        ws.cell(row=row, column=4).value = to_account  # D 转入账户
        applied += 1

    out_path = args.output or args.merged_xlsx.replace(".xlsx", "_matched.xlsx")
    wb.save(out_path)
    wb.close()
    print(f"应用完成：改写 {applied} 条，跳过 {skipped} 条 → {out_path}")


def main():
    parser = argparse.ArgumentParser(description="LLM 模糊匹配桥梁")
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_extract = sub.add_parser("extract", help="提取未匹配的转账/亲属卡")
    p_extract.add_argument("merged_xlsx")
    p_extract.add_argument("tasks_json")
    p_extract.set_defaults(func=cmd_extract)

    p_apply = sub.add_parser("apply", help="把 LLM 匹配结果回写到 xlsx")
    p_apply.add_argument("merged_xlsx")
    p_apply.add_argument("results_json")
    p_apply.add_argument("--output", "-o")
    p_apply.set_defaults(func=cmd_apply)

    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
