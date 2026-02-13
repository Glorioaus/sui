"""
合并处理脚本
读取所有Excel文件，执行跨文件退款对冲和转账识别
"""
import os
import sys
import re
from datetime import datetime, timedelta
from typing import List, Tuple, Optional
import openpyxl

# 添加src目录到路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from models import Transaction
from excel_generator import ExcelGenerator


# 转账识别关键词映射
TRANSFER_KEYWORDS = {
    "中信": "中信信用卡",
    "招商": "招商信用卡",
    "浦发": "浦发信用卡",
    "建行信用": "建行信用卡",
    "建行卡": "建行信用卡",
    "花呗": "花呗",
    "京东白条": "京东白条",
    "微信": "微信",
    "零钱": "微信",
    "支付宝": "支付宝",
    "余额宝": "支付宝",
    "信用卡还款": None,  # 通用信用卡还款
    "还款": None,
    "跨行还款": None,  # 需要通过金额匹配确定目标
}

# 储蓄卡账户列表（转账来源）
DEBIT_ACCOUNTS = ["农业银行", "宁波银行", "建行储蓄卡", "工商储蓄卡", "招商储蓄卡", "农村信用合作社"]

# 信用卡/钱包账户列表（转账目标）
CREDIT_ACCOUNTS = ["中信信用卡", "浦发信用卡", "招商信用卡", "建行信用卡", "信用卡", "花呗", "京东白条", "微信", "支付宝"]


def read_excel_transactions(file_path: str) -> List[Transaction]:
    """
    从Excel文件读取交易记录
    支持新格式（3个Sheet：支出、收入、转账）
    """
    transactions = []
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)

        # 检查是否是新格式（有支出/收入/转账 sheet）
        sheet_names = wb.sheetnames
        is_new_format = "支出" in sheet_names or "收入" in sheet_names or "转账" in sheet_names

        if is_new_format:
            # 新格式：读取3个sheet
            transactions.extend(_read_expense_sheet(wb, sheet_names))
            transactions.extend(_read_income_sheet(wb, sheet_names))
            transactions.extend(_read_transfer_sheet(wb, sheet_names))
        else:
            # 旧格式：读取单个活动sheet
            transactions.extend(_read_legacy_sheet(wb))

        wb.close()
    except Exception as e:
        print(f"读取文件失败 {file_path}: {e}")

    return transactions


def _read_expense_sheet(wb, sheet_names: List[str]) -> List[Transaction]:
    """读取支出sheet"""
    transactions = []
    if "支出" not in sheet_names:
        return transactions

    ws = wb["支出"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue

        # 支出格式: 交易类型(0), 日期(1), 分类(2), 子分类(3), 支出账户(4), 金额(5), 成员(6), 商家(7), 项目(8), 备注(9)
        date_val = row[1]
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val).split()[0] if date_val else ""

        amount = float(row[5]) if row[5] else 0.0

        transaction = Transaction(
            date=date_str,
            category=str(row[2] or ""),
            subcategory=str(row[3] or ""),
            account=str(row[4] or ""),
            amount=amount,
            description=str(row[9] or "") if len(row) > 9 else "",
            transaction_type="支出",
            merchant=str(row[7] or "") if len(row) > 7 else "",
        )
        transactions.append(transaction)

    return transactions


def _read_income_sheet(wb, sheet_names: List[str]) -> List[Transaction]:
    """读取收入sheet"""
    transactions = []
    if "收入" not in sheet_names:
        return transactions

    ws = wb["收入"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue

        # 收入格式: 交易类型(0), 日期(1), 分类(2), 子分类(3), 收入账户(4), 金额(5), 成员(6), 商家(7), 项目(8), 备注(9)
        date_val = row[1]
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val).split()[0] if date_val else ""

        amount = float(row[5]) if row[5] else 0.0

        transaction = Transaction(
            date=date_str,
            category=str(row[2] or ""),
            subcategory=str(row[3] or ""),
            account=str(row[4] or ""),
            amount=amount,
            description=str(row[9] or "") if len(row) > 9 else "",
            transaction_type="收入",
            merchant=str(row[7] or "") if len(row) > 7 else "",
        )
        transactions.append(transaction)

    return transactions


def _read_transfer_sheet(wb, sheet_names: List[str]) -> List[Transaction]:
    """读取转账sheet"""
    transactions = []
    if "转账" not in sheet_names:
        return transactions

    ws = wb["转账"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue

        # 转账格式: 交易类型(0), 日期(1), 转出账户(2), 转入账户(3), 金额(4), 成员(5), 商家(6), 项目(7), 备注(8)
        date_val = row[1]
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val).split()[0] if date_val else ""

        amount = float(row[4]) if row[4] else 0.0

        transaction = Transaction(
            date=date_str,
            category="转账",
            subcategory="",
            account=str(row[2] or ""),  # 转出账户
            amount=amount,
            description=str(row[8] or "") if len(row) > 8 else "",
            transaction_type="转账",
            transfer_to_account=str(row[3] or ""),  # 转入账户
            merchant=str(row[6] or "") if len(row) > 6 else "",
        )
        transactions.append(transaction)

    return transactions


def _read_legacy_sheet(wb) -> List[Transaction]:
    """读取旧格式的单个sheet"""
    transactions = []
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue

        # 旧格式: 交易日期(0), 分类(1), 类型(2), 子分类(3), 支付账户(4), 金额(5), 成员(6), 商家(7), 项目(8), 备注(9)
        date_val = row[0]
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val)

        amount = float(row[5]) if row[5] else 0.0

        transaction = Transaction(
            date=date_str,
            category=str(row[1] or ""),
            subcategory=str(row[3] or ""),
            account=str(row[4] or ""),
            amount=amount,
            description=str(row[9] or ""),
            transaction_type=str(row[2] or "支出"),
            merchant=str(row[7] or ""),
        )
        transactions.append(transaction)

    return transactions


def parse_date(date_str: str) -> Optional[datetime]:
    """解析日期字符串"""
    formats = ["%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日"]
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None


def dates_within_range(date1: str, date2: str, days: int = 3) -> bool:
    """检查两个日期是否在指定天数范围内"""
    d1 = parse_date(date1)
    d2 = parse_date(date2)
    if d1 is None or d2 is None:
        return False
    return abs((d1 - d2).days) <= days


def amounts_match(amount1: float, amount2: float, tolerance: float = 0.01) -> bool:
    """检查两个金额是否匹配（允许小误差）"""
    return abs(amount1 - amount2) <= tolerance


def normalize_merchant(merchant: str, description: str) -> str:
    """
    标准化商户名称，用于匹配
    从商户名或描述中提取关键信息
    """
    # 优先使用商户名
    text = merchant if merchant else description
    if not text:
        return ""

    # 移除常见前缀后缀
    text = re.sub(r'^(消费|支付|转账|退款|退货)[:\-\s]*', '', text)
    text = re.sub(r'[\s\-_]+', '', text)

    return text.strip().lower()


def is_masked_or_person_name(merchant: str) -> bool:
    """
    检查商户名是否是脱敏的或人名
    脱敏格式如：天猫**营、淘宝**店
    人名通常是2-4个汉字
    """
    if not merchant:
        return False
    # 脱敏商户名（含**）
    if "**" in merchant:
        return True
    # 短名称可能是人名（2-4个汉字）
    if len(merchant) <= 4 and all('\u4e00' <= c <= '\u9fff' for c in merchant):
        return True
    return False


def reconcile_refunds(transactions: List[Transaction]) -> List[Transaction]:
    """
    执行退款对冲
    第一轮：同商户 + 同金额（精确匹配）
    第二轮：脱敏/人名商户 + 同金额 + 日期接近（模糊匹配）
    结果：匹配成功的消费和退款都删除
    """
    print("\n=== 开始退款对冲 ===")

    # 分离支出和收入（退款）
    expenses = []  # (index, transaction)
    refunds = []   # (index, transaction)

    for i, t in enumerate(transactions):
        if t.transaction_type == "支出":
            expenses.append((i, t))
        elif t.transaction_type == "收入":
            # 检查是否是退款类型
            desc_lower = (t.description or "").lower()
            cat_lower = (t.category or "").lower()
            subcat_lower = (t.subcategory or "").lower()
            if "退款" in desc_lower or "退款" in cat_lower or "退款" in subcat_lower:
                refunds.append((i, t))

    # 记录要删除的索引
    to_remove = set()
    matched_count = 0
    fuzzy_matched_count = 0

    # 第一轮：精确商户匹配
    for ref_idx, refund in refunds:
        if ref_idx in to_remove:
            continue

        ref_merchant = normalize_merchant(refund.merchant, refund.description)

        for exp_idx, expense in expenses:
            if exp_idx in to_remove:
                continue

            exp_merchant = normalize_merchant(expense.merchant, expense.description)

            # 匹配条件：商户相似 + 金额相同
            if ref_merchant and exp_merchant and ref_merchant == exp_merchant:
                if amounts_match(refund.amount, expense.amount):
                    desc = (expense.merchant or expense.description[:20]).encode('gbk', errors='replace').decode('gbk')
                    print(f"  精确匹配: [{expense.date}] {desc} "
                          f"{expense.amount} <-> [{refund.date}] 退款 {refund.amount}")
                    to_remove.add(ref_idx)
                    to_remove.add(exp_idx)
                    matched_count += 1
                    break

    # 第二轮：模糊匹配（针对脱敏商户名或人名）
    for ref_idx, refund in refunds:
        if ref_idx in to_remove:
            continue

        ref_merchant = refund.merchant or ""

        # 只对脱敏商户名或人名进行模糊匹配
        if not is_masked_or_person_name(ref_merchant):
            continue

        # 按金额匹配 + 日期接近（±30天，因为退款可能很晚）
        for exp_idx, expense in expenses:
            if exp_idx in to_remove:
                continue

            if amounts_match(refund.amount, expense.amount):
                if dates_within_range(refund.date, expense.date, days=30):
                    desc = (expense.merchant or expense.description[:20]).encode('gbk', errors='replace').decode('gbk')
                    print(f"  模糊匹配: [{expense.date}] {desc} "
                          f"{expense.amount} <-> [{refund.date}] 退款({ref_merchant}) {refund.amount}")
                    to_remove.add(ref_idx)
                    to_remove.add(exp_idx)
                    fuzzy_matched_count += 1
                    break

    # 过滤掉已对冲的记录
    result = [t for i, t in enumerate(transactions) if i not in to_remove]

    print(f"退款对冲完成：精确匹配 {matched_count} 对，模糊匹配 {fuzzy_matched_count} 对")
    print(f"  删除 {len(to_remove)} 条记录")
    return result


def identify_transfer_target(description: str) -> Optional[str]:
    """
    从描述中识别转账目标账户
    """
    if not description:
        return None

    desc_lower = description.lower()

    for keyword, target in TRANSFER_KEYWORDS.items():
        if keyword.lower() in desc_lower:
            return target

    return None


def identify_transfers(transactions: List[Transaction]) -> List[Transaction]:
    """
    执行转账识别
    匹配条件：
    - 账户A有"支出"（如农行转出）
    - 账户B有"收入"或被识别为还款
    - 金额相同，日期接近（±3天）
    结果：删除两条记录，生成一条转账记录

    新增：对于"跨行还款"等无明确目标的记录，通过金额+日期匹配信用卡还款记录来确定目标
    """
    print("\n=== 开始转账识别 ===")

    # 分离储蓄卡支出（可能是转账）
    debit_expenses_with_target = []  # (index, transaction, target) - 有明确目标
    debit_expenses_need_match = []   # (index, transaction) - 需要通过匹配确定目标
    credit_incomes = []  # (index, transaction)

    for i, t in enumerate(transactions):
        # 储蓄卡支出
        if t.account in DEBIT_ACCOUNTS and t.transaction_type == "支出":
            # 跳过已被分类为按揭还款的交易（避免误识别为信用卡转账）
            if t.category == "金融保险" and t.subcategory == "按揭还款":
                continue

            target = identify_transfer_target(t.description)
            if target:
                # 有明确目标（如"中信" → 中信信用卡）
                debit_expenses_with_target.append((i, t, target))
            else:
                # 检查是否含还款关键词但无明确目标
                desc = (t.description or "").lower()
                if "跨行还款" in desc or "还款" in desc or "信用卡" in desc:
                    debit_expenses_need_match.append((i, t))

        # 信用卡收入（还款）- 包括普通收入和特殊标记的还款记录
        if t.account in CREDIT_ACCOUNTS and t.transaction_type == "收入":
            credit_incomes.append((i, t))
        # 特殊还款标记（来自信用卡解析器，用于匹配后删除）
        elif t.category == "__REPAYMENT__" and t.transaction_type == "收入":
            credit_incomes.append((i, t))

    # 记录要删除的索引和新增的转账记录
    to_remove = set()
    transfers = []
    matched_count = 0

    # 第一轮：处理有明确目标的转账
    for exp_idx, expense, target in debit_expenses_with_target:
        if exp_idx in to_remove:
            continue

        matched_income = False

        # 尝试匹配信用卡收入
        for inc_idx, income in credit_incomes:
            if inc_idx in to_remove:
                continue

            # 检查是否是目标账户
            if income.account != target:
                continue

            # 检查金额和日期
            if amounts_match(expense.amount, income.amount) and \
               dates_within_range(expense.date, income.date):
                print(f"  转账匹配: [{expense.date}] {expense.account} -> {income.account} "
                      f"{expense.amount}")

                # 确定子分类：支付宝/微信用"充值"，其他用"还款"
                subcategory = "充值" if income.account in ["支付宝", "微信"] else "还款"

                # 创建转账记录
                transfer = Transaction(
                    date=expense.date,
                    category="转账",
                    subcategory=subcategory,
                    account=expense.account,
                    amount=expense.amount,
                    description=expense.description,
                    transaction_type="转账",
                    transfer_to_account=income.account,
                )
                transfers.append(transfer)

                to_remove.add(exp_idx)
                to_remove.add(inc_idx)
                matched_count += 1
                matched_income = True
                break

        # 如果没有匹配到信用卡收入，但有明确目标，仍标记为转账
        if not matched_income and exp_idx not in to_remove:
            print(f"  转账标记: [{expense.date}] {expense.account} -> {target} "
                  f"{expense.amount}")

            subcategory = "充值" if target in ["支付宝", "微信"] else "还款"

            transfer = Transaction(
                date=expense.date,
                category="转账",
                subcategory=subcategory,
                account=expense.account,
                amount=expense.amount,
                description=expense.description,
                transaction_type="转账",
                transfer_to_account=target,
            )
            transfers.append(transfer)

            to_remove.add(exp_idx)
            matched_count += 1

    # 第二轮：处理需要通过金额匹配确定目标的转账（如"跨行还款"）
    for exp_idx, expense in debit_expenses_need_match:
        if exp_idx in to_remove:
            continue

        matched = False

        # 遍历所有信用卡收入，按金额+日期匹配
        for inc_idx, income in credit_incomes:
            if inc_idx in to_remove:
                continue

            # 检查金额和日期
            if amounts_match(expense.amount, income.amount) and \
               dates_within_range(expense.date, income.date):
                print(f"  跨行还款匹配: [{expense.date}] {expense.account} -> {income.account} "
                      f"{expense.amount} (通过金额匹配)")

                subcategory = "充值" if income.account in ["支付宝", "微信"] else "还款"

                transfer = Transaction(
                    date=expense.date,
                    category="转账",
                    subcategory=subcategory,
                    account=expense.account,
                    amount=expense.amount,
                    description=expense.description,
                    transaction_type="转账",
                    transfer_to_account=income.account,
                )
                transfers.append(transfer)

                to_remove.add(exp_idx)
                to_remove.add(inc_idx)
                matched_count += 1
                matched = True
                break

        # 如果无法匹配但确实含有还款关键词，仍标记为转账到"信用卡"
        if not matched:
            print(f"  跨行还款(未匹配): [{expense.date}] {expense.account} -> 信用卡 "
                  f"{expense.amount} (无法确定具体卡)")

            transfer = Transaction(
                date=expense.date,
                category="转账",
                subcategory="还款",
                account=expense.account,
                amount=expense.amount,
                description=expense.description,
                transaction_type="转账",
                transfer_to_account="信用卡",
            )
            transfers.append(transfer)

            to_remove.add(exp_idx)
            matched_count += 1

    # 过滤并添加转账记录
    # 同时删除未匹配的 __REPAYMENT__ 标记记录（它们只是用于匹配的临时记录）
    result = [t for i, t in enumerate(transactions)
              if i not in to_remove and t.category != "__REPAYMENT__"]
    result.extend(transfers)

    print(f"转账识别完成：{matched_count} 条识别，删除 {len(to_remove)} 条原记录")
    return result


def process_family_card(transactions: List[Transaction]) -> List[Transaction]:
    """
    处理亲属卡/亲友代付交易
    将微信"亲属卡交易"和支付宝"亲友代付"对应的银行卡支出重分类为"其他杂项-XX支出"
    """
    print("\n=== 开始亲属卡处理 ===")

    # 找出所有标记交易（来自微信/支付宝的亲属卡标记）
    markers = []  # (index, transaction, user_name, target_bank)
    for i, t in enumerate(transactions):
        if t.category == "__FAMILY_CARD__" or t.transaction_type == "__MARKER__":
            user_name = t.subcategory or "亲属"  # 使用者名称存在subcategory中
            target_bank = t.account  # 目标银行（可能是具体银行名或"__ANY_BANK__"）
            markers.append((i, t, user_name, target_bank))

    if not markers:
        print("未发现亲属卡标记")
        return transactions

    # 所有银行卡账户（用于匹配）
    all_bank_accounts = set(DEBIT_ACCOUNTS + CREDIT_ACCOUNTS)

    # 统计各银行账户在数据中是否有交易
    accounts_with_data = set()
    for t in transactions:
        if t.category != "__FAMILY_CARD__" and t.account in all_bank_accounts:
            accounts_with_data.add(t.account)

    print(f"  数据中存在的银行账户: {', '.join(sorted(accounts_with_data))}")

    # 记录要删除的标记索引
    marker_indices_to_remove = set()
    matched_tx_indices = set()
    matched_count = 0
    unmatched_deleted_count = 0  # 未匹配但银行有数据（删除避免重复）
    unmatched_kept_count = 0     # 未匹配且银行无数据（保留）

    for marker_idx, marker, user_name, target_bank in markers:
        found_match = False

        # 在银行卡交易中查找匹配的交易
        for i, t in enumerate(transactions):
            if i in marker_indices_to_remove or i in matched_tx_indices:
                continue
            if t.category == "__FAMILY_CARD__":
                continue
            if t.account not in all_bank_accounts or t.account == "微信":
                continue

            # 匹配条件：日期 + 金额
            if t.date == marker.date and amounts_match(t.amount, marker.amount):
                # 如果指定了具体银行，还要匹配账户
                if target_bank != "__ANY_BANK__" and t.account != target_bank:
                    continue

                print(f"  亲属卡匹配: [{t.date}] {t.account} {t.amount} -> {user_name}支出")

                # 重分类为"其他杂项-XX支出"
                t.category = "其他杂项"
                t.subcategory = f"{user_name}支出"
                t.transaction_type = "支出"
                matched_tx_indices.add(i)
                marker_indices_to_remove.add(marker_idx)
                matched_count += 1
                found_match = True
                break

        # 未匹配的处理
        if not found_match:
            # 检查目标银行是否有数据
            bank_has_data = (target_bank in accounts_with_data) or (target_bank == "__ANY_BANK__")

            if bank_has_data:
                # 银行有数据但未匹配 → 删除标记（避免重复）
                marker_indices_to_remove.add(marker_idx)
                unmatched_deleted_count += 1
            else:
                # 银行无数据 → 保留标记作为唯一记录
                marker.category = "其他杂项"
                marker.subcategory = f"{user_name}支出"
                marker.transaction_type = "支出"
                if "微信" in (marker.description or ""):
                    marker.account = "微信"
                else:
                    marker.account = "支付宝"
                unmatched_kept_count += 1

    # 过滤掉需要删除的标记
    result = [t for i, t in enumerate(transactions) if i not in marker_indices_to_remove]

    print(f"亲属卡处理完成：")
    print(f"  匹配成功: {matched_count} 条（重分类银行卡交易）")
    print(f"  未匹配-删除: {unmatched_deleted_count} 条（银行有数据，避免重复）")
    print(f"  未匹配-保留: {unmatched_kept_count} 条（银行无数据）")
    return result


def sort_transactions(transactions: List[Transaction]) -> List[Transaction]:
    """按日期排序交易记录"""
    def sort_key(t):
        d = parse_date(t.date)
        return d if d else datetime.min

    return sorted(transactions, key=sort_key)


def merge_excel_files(input_dir: str, output_path: str = None):
    """
    合并处理主函数
    """
    print(f"=== 开始合并处理 ===")
    print(f"输入目录: {input_dir}")

    # 查找所有Excel文件
    excel_files = []
    for f in os.listdir(input_dir):
        if f.endswith('.xlsx') and not f.startswith('~') and not f.startswith('merged'):
            excel_files.append(os.path.join(input_dir, f))

    if not excel_files:
        print("未找到Excel文件")
        return

    print(f"找到 {len(excel_files)} 个Excel文件")

    # 读取所有交易记录
    all_transactions = []
    for file_path in excel_files:
        print(f"  读取: {os.path.basename(file_path)}")
        transactions = read_excel_transactions(file_path)
        all_transactions.extend(transactions)
        print(f"    {len(transactions)} 条记录")

    print(f"\n合计 {len(all_transactions)} 条交易记录")

    # 执行退款对冲
    transactions = reconcile_refunds(all_transactions)

    # 执行转账识别
    transactions = identify_transfers(transactions)

    # 执行亲属卡/亲友代付处理
    transactions = process_family_card(transactions)

    # 按日期排序
    transactions = sort_transactions(transactions)

    # 生成输出文件
    if output_path is None:
        output_path = os.path.join(input_dir, "merged_账单.xlsx")

    print(f"\n=== 生成合并文件 ===")
    generator = ExcelGenerator()
    generator._create_workbook()
    generator.add_transactions(transactions)
    generator.save(output_path)

    print(f"\n处理完成！")
    print(f"  输入文件: {len(excel_files)} 个")
    print(f"  原始记录: {len(all_transactions)} 条")
    print(f"  最终记录: {len(transactions)} 条")
    print(f"  输出文件: {output_path}")


def main():
    if len(sys.argv) < 2:
        print("用法: python merge.py <输出目录> [合并文件名]")
        print("示例: python merge.py output/")
        print("      python merge.py output/ merged.xlsx")
        sys.exit(1)

    input_dir = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None

    if not os.path.isdir(input_dir):
        print(f"错误: 目录不存在 {input_dir}")
        sys.exit(1)

    merge_excel_files(input_dir, output_path)


if __name__ == "__main__":
    main()
