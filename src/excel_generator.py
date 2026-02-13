"""
Excel生成器模块
将账单数据生成随手记可导入的Excel文件
支持三个分页：支出、收入、转账
"""
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from typing import List, Dict
from models import Transaction, BankStatement


class ExcelGenerator:
    """
    Excel生成器类
    生成随手记兼容的Excel格式（3个Sheet）
    """

    # 支出Sheet列定义
    EXPENSE_COLUMNS = [
        "交易类型",  # A
        "日期",      # B
        "分类",      # C
        "子分类",    # D
        "支出账户",  # E
        "金额",      # F
        "成员",      # G
        "商家",      # H
        "项目",      # I
        "备注"       # J
    ]

    # 收入Sheet列定义
    INCOME_COLUMNS = [
        "交易类型",  # A
        "日期",      # B
        "分类",      # C
        "子分类",    # D
        "收入账户",  # E
        "金额",      # F
        "成员",      # G
        "商家",      # H
        "项目",      # I
        "备注"       # J
    ]

    # 转账Sheet列定义
    TRANSFER_COLUMNS = [
        "交易类型",  # A
        "日期",      # B
        "转出账户",  # C
        "转入账户",  # D
        "金额",      # E
        "成员",      # F
        "商家",      # G
        "项目",      # H
        "备注"       # I
    ]

    def __init__(self, template_path: str = None):
        """
        初始化Excel生成器
        """
        self.template_path = template_path
        self.workbook = None
        self.sheets: Dict[str, openpyxl.worksheet.worksheet.Worksheet] = {}
        self.sheet_rows: Dict[str, int] = {}

    def _create_workbook(self):
        """
        创建新的工作簿并设置三个Sheet
        """
        self.workbook = openpyxl.Workbook()

        # 创建三个Sheet
        # 删除默认sheet
        default_sheet = self.workbook.active
        self.workbook.remove(default_sheet)

        # 按顺序创建：支出、收入、转账
        self.sheets["支出"] = self.workbook.create_sheet("支出")
        self.sheets["收入"] = self.workbook.create_sheet("收入")
        self.sheets["转账"] = self.workbook.create_sheet("转账")

        # 初始化每个sheet的当前行
        self.sheet_rows = {"支出": 2, "收入": 2, "转账": 2}

        # 设置表头
        self._setup_sheet_header(self.sheets["支出"], self.EXPENSE_COLUMNS)
        self._setup_sheet_header(self.sheets["收入"], self.INCOME_COLUMNS)
        self._setup_sheet_header(self.sheets["转账"], self.TRANSFER_COLUMNS)

        print("创建新工作簿完成（支出、收入、转账三个Sheet）")

    def _setup_sheet_header(self, worksheet, columns: List[str]):
        """
        设置Sheet表头
        """
        # 设置表头样式
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 写入表头
        for col, header in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 设置列宽
        if len(columns) == 10:  # 支出/收入
            column_widths = [10, 20, 12, 12, 12, 10, 8, 15, 10, 30]
        else:  # 转账 (9列)
            column_widths = [10, 20, 12, 12, 10, 8, 15, 10, 30]

        for col, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    def add_transaction(self, transaction: Transaction):
        """
        添加交易记录到对应的Sheet
        """
        if not self.sheets:
            raise Exception("请先创建工作簿")

        tx_type = transaction.transaction_type

        if tx_type == "支出":
            self._add_expense(transaction)
        elif tx_type == "收入":
            self._add_income(transaction)
        elif tx_type == "转账":
            self._add_transfer(transaction)
        else:
            # 默认归类为支出
            self._add_expense(transaction)

    def _add_expense(self, transaction: Transaction):
        """
        添加支出记录
        """
        ws = self.sheets["支出"]
        row = self.sheet_rows["支出"]

        # 交易类型, 日期, 分类, 子分类, 支出账户, 金额, 成员, 商家, 项目, 备注
        ws.cell(row=row, column=1, value="支出")
        ws.cell(row=row, column=2, value=transaction.date)
        ws.cell(row=row, column=3, value=transaction.category)
        ws.cell(row=row, column=4, value=transaction.subcategory)
        ws.cell(row=row, column=5, value=transaction.account)
        ws.cell(row=row, column=6, value=transaction.amount)
        ws.cell(row=row, column=7, value="")  # 成员
        ws.cell(row=row, column=8, value=getattr(transaction, 'merchant', None) or "")
        ws.cell(row=row, column=9, value="")  # 项目
        ws.cell(row=row, column=10, value=transaction.description or "")

        self.sheet_rows["支出"] += 1

    def _add_income(self, transaction: Transaction):
        """
        添加收入记录
        """
        ws = self.sheets["收入"]
        row = self.sheet_rows["收入"]

        # 交易类型, 日期, 分类, 子分类, 收入账户, 金额, 成员, 商家, 项目, 备注
        ws.cell(row=row, column=1, value="收入")
        ws.cell(row=row, column=2, value=transaction.date)
        ws.cell(row=row, column=3, value=transaction.category)
        ws.cell(row=row, column=4, value=transaction.subcategory)
        ws.cell(row=row, column=5, value=transaction.account)
        ws.cell(row=row, column=6, value=transaction.amount)
        ws.cell(row=row, column=7, value="")  # 成员
        ws.cell(row=row, column=8, value=getattr(transaction, 'merchant', None) or "")
        ws.cell(row=row, column=9, value="")  # 项目
        ws.cell(row=row, column=10, value=transaction.description or "")

        self.sheet_rows["收入"] += 1

    def _add_transfer(self, transaction: Transaction):
        """
        添加转账记录
        """
        ws = self.sheets["转账"]
        row = self.sheet_rows["转账"]

        # 交易类型, 日期, 转出账户, 转入账户, 金额, 成员, 商家, 项目, 备注
        ws.cell(row=row, column=1, value="转账")
        ws.cell(row=row, column=2, value=transaction.date)
        ws.cell(row=row, column=3, value=transaction.account)  # 转出账户
        ws.cell(row=row, column=4, value=transaction.transfer_to_account or "")  # 转入账户
        ws.cell(row=row, column=5, value=transaction.amount)
        ws.cell(row=row, column=6, value="")  # 成员
        ws.cell(row=row, column=7, value=getattr(transaction, 'merchant', None) or "")
        ws.cell(row=row, column=8, value="")  # 项目
        ws.cell(row=row, column=9, value=transaction.description or "")

        self.sheet_rows["转账"] += 1

    def add_transactions(self, transactions: List[Transaction]):
        """
        批量添加交易记录
        """
        for transaction in transactions:
            self.add_transaction(transaction)

    def save(self, output_path: str):
        """
        保存Excel文件
        """
        if self.workbook is None:
            raise Exception("没有工作簿可保存")

        # 确保输出目录存在
        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)

        try:
            self.workbook.save(output_path)
            print(f"成功保存文件：{output_path}")
        except Exception as e:
            raise Exception(f"保存文件失败：{e}")

    def generate(self, statement: BankStatement, output_path: str):
        """
        生成随手记Excel文件
        """
        self._create_workbook()
        self.add_transactions(statement.transactions)
        self.save(output_path)

        # 统计各类型数量
        expense_count = self.sheet_rows["支出"] - 2
        income_count = self.sheet_rows["收入"] - 2
        transfer_count = self.sheet_rows["转账"] - 2

        print(f"生成完成，共 {len(statement.transactions)} 条记录")
        print(f"  支出: {expense_count} 条")
        print(f"  收入: {income_count} 条")
        print(f"  转账: {transfer_count} 条")
